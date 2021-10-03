#coding: utf-8

from datetime import timedelta, date
import json

# Cette recette lit une feuille excel et produit un fichier de json de config 
# contanant les information sur le programme de garde des groupes de pharmacies
# l'entete doit respecter ce format : ['debut_periode', 'fin_periode', 'groupe']
# changer les variables file_path et sheetname puis excecuter


import openpyxl
import json

# Chemin du fichier excel contenant le programme
file_path = '../data/Programme de garde 2021.xlsx'
# nom de la feuille
sheetname = '2021'

def daterange(date1, date2):
    for n in range(int ((date2 - date1).days)):
        yield date1 + timedelta(n)

excel_file = openpyxl.load_workbook(file_path)
sheet = excel_file[sheetname]
max_row = sheet.max_row
max_col = sheet.max_column
min_col = sheet.min_column
min_row = sheet.min_row
output_config_dic = {}
    
header = [str((sheet.cell(row = min_row, column = i).value)) for i in range(1, max_col+1)]
for h in header:
    if h not in ['debut_periode', 'fin_periode', 'groupe']:
        raise NameError

for r in range(min_row+1, max_row+1):
    start_date = sheet.cell(row = r, column = min_col).value
    end_date = sheet.cell(row = r, column = min_col+1).value
    group = sheet.cell(row = r, column = min_col+2).value
    for dt in daterange(start_date, end_date):
        key = dt.strftime("%Y-%m-%d")
        value = group
        output_config_dic[key] = value

with open(r'../pharmacies/pharmacies_programm.json', 'w') as outfile:
    json.dump(output_config_dic, outfile, indent=2)
    


