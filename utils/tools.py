#coding: utf-8

import openpyxl
import json

def exel_to_json(file_path: str, sheetname: str, replace_none_value=None, lower=True):
    """ convert excel sheet to dict. the first ligne is the header """
    
    excel_file = openpyxl.load_workbook(file_path)
    sheet = excel_file[sheetname]
    max_row = sheet.max_row
    max_col = sheet.max_column
    min_col = sheet.min_column
    min_row = sheet.min_row
    json_items_list = []
    
    header = [str((sheet.cell(row = min_row, column = i).value)) for i in range(1, max_col+1)]
    
    tmp_list = []
    for r in range(min_row+1, max_row+1):
        for c in range(min_col, max_col+1):
            cell_value = sheet.cell(row = r, column = c).value
            if replace_none_value is not None and cell_value is None:
                cell_value = replace_none_value
            if lower is True and type(cell_value) is str:
                cell_value = cell_value.lower().strip()
            tmp_list.append(cell_value)
        obj = dict(zip(header, tmp_list))
        if all(value == None for value in obj.values()) == False:   
            json_items_list.append(obj)
        tmp_list = []
    # fix encoded problem
    data = json.dumps(json_items_list, indent=4, ensure_ascii=False)
    return json.loads(data)
    
def split_values(list_json_obj: list, fields_name:list) -> list: 
    """split values ​​contained in fields of a list of json objects
        example:
            d = [{'last_name': 'Abdoul Fataoh'}, {'last_name': ' Djami Farida'} ]
            field = [{'key': 'last_name', 'separator':' '}]
            split_values(d, field)
            --> [{'last_name': ['Abdoul', 'Fataoh']}, {'last_name': ['Djami', 'Farida']}]

    Args:
        list_json_obj (list): list of json objects. objects must have the same keys
        fields_name (list): list of selected fields to split

    Returns:
        list: output list
    """
    for i in range(len(list_json_obj)):
        for p in range(len(fields_name)):
            try:
                t = list_json_obj[i][fields_name[p]['key']].split(fields_name[p]['separator'])
            except AttributeError:
                    t = []
            list_json_obj[i][fields_name[p]['key']] = t
    return list_json_obj

# Test
if __name__ == '__main__':
    d = exel_to_json('../data/groupes_pharmacies.xlsx', 'groupe_4')
    # d = split_values(d,  [{'key':'dci', 'separator': ','}])
    print(d)
